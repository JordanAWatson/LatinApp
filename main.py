# This program assists in the memorization of latin verbs through repetition
#
# Author: Jordan Watson
# Date: 3/10/2021
# Last Updated: 9/13/2021

# TODO
# I'm not going to remeber anything without more comments
# Add an irregular verb section
# Add support for deponent and semi-deponent verbs
# implement noun practice if I'm feeling crazy
# stop trusting user input


# this function clears the screen
# taken from https://stackoverflow.com/questions/517970/how-to-clear-the-interpreter-console
import os
def cls():
    os.system('cls' if os.name=='nt' else 'clear')

# This is the main loop for entering the practise section
def verbs():
    from openpyxl import load_workbook
    FILE_NAME = 'verbs.xlsx'
    wb = load_workbook(FILE_NAME, read_only = True)

    practise(wb)  # I really hate how the word practise looks

    while True:
        choice = input('Practise something else? y/n')
        if choice == 'y' or choice == 'Y':
            practise(wb)
        else:
            break;

    # workbook is read only and must be closed manually
    wb.close()

# This function is a single iteration of a practise cycle
# takes user input and compares that to an answer key
def practise(wb):
    import random
    cls()
    print("+=============================================================+")
    print('What conjugation do you want to practice?')
    print(wb.sheetnames)
    conj = input()

    ws = wb[conj]

    # counts the number of rows
    numRows = 0;
    for x in ws.values:
        numRows += 1

    cls()
    print("+=============================================================+")
    print(numRows - 1, 'verbs in', conj)


    print('What tense would you like to practice in?')
    print('1) Indicative Active')
    print('2) Subjunctive Active')
    print('3) Indicative Passive')
    print('4) Subjunctive Passive')
    print('5) Imperative')
    tense = int(input())

    index = 0;
    if tense == 1:
        print('1) Present')
        print('2) Imperfect')
        print('3) Future')
        print('4) Perfect')
        print('5) Pluperfect')
        print('6) Future Perfect')
        index = int(input()) - 1

    elif tense == 2:
        print('1) Present')
        print('2) Imperfect')
        print('3) Perfect')
        print('4) Pluperfect')
        index = int(input()) + 5

    elif tense == 3:
        print('1) Present')
        print('2) Imperfect')
        print('3) Future')
        print('4) Perfect')
        print('5) Pluperfect')
        print('6) Future Perfect')
        index = int(input()) + 9

    elif tense == 4:
        print('1) Present')
        print('2) Imperfect')
        index = int(input()) + 11

    elif tense == 5:
        print('1) Present Active')
        print('2) Present Passive')
        index = int(input()) + 13

    else:
        print('wrong input')

    # change these for excel file formatting
    index *= 6  # parts per tense
    index += 6  # principle parts

    random.seed(a=None, version=2)

    print('How many times do you want to practise?')
    repeat = int(input())
    numCorrect = 0
    
    for i in range(repeat):
        # pick a random row
        r = random.randint(2, numRows)

        # print principle parts
        cls()
        print("+=============================================================+")
        print('Principle parts:')
        for x in range(5):
            c = ws.cell(row = r, column = (x + 1))
            print(c.value, end = ' ')
            if x == 3:
                print()
        print()

        # There has to be a better way
        answer = ['a', 'a', 'a', 'a', 'a', 'a']
        key = ['a', 'a', 'a', 'a', 'a', 'a']

        # get user answers, compare them with key, then show incorrect
        print('One at a time')
        for x in range(6):
            c = ws.cell(row = r, column = index + x)
            key[x] = str(c.value)
            answer[x] = str(input())

        # don't want to give the answer right away
        correct = True
        for x in range(6):
            if key[x] != answer[x]:
                print('Given:', answer[x], ' Expected:', key[x])
                correct = False

        if correct == True:
            print("Correct")
            numCorrect += 1

        input('Press "enter" to continue')
        cls()

    print('Finished Practice')
    print('You got', numCorrect, '/', repeat,
          "correct, Nice Job!" if numCorrect == repeat else "correct")
    print()
    # End practise()

# this function prints a how to use guide
def how_to():
    cls()
    print("+=============================================================+")
    print("This practise system is based on the Dowling Method")
    print("Learn about it here: https://wcdrutgers.net/Latin.htm")
    print("or Google (or whatever search engine you use) 'Dowling Latin'")
    print()
    print("+=============================================================+")
    print("How to enter verbs:")
    print("All verb conjugations should be entered like this,")
    print("1st")
    print("2nd")
    print("3rd")
    print("1st plural")
    print("2nd plural")
    print("3rd plural")
    print()
    print("So after typing each form, press Enter")
    print("Afterwards you will be told wich ones, if any, you got wrong.")
    print()
    print("+=============================================================+")
    print("Macrons, those bars above the vowels:")
    print("Yes I do require macrons. They are an important part of the")
    print("language, and a lack of macrons in other practice programs is")
    print("why I created this in the first place.")
    print("To use macrons on windows set your keyboard language to Maori,")
    print("then press the tilde [~`] button before any long vowel.")
    print("+=============================================================+")
    input("Press 'Enter' for the next page")
    cls()
    print("+=============================================================+")
    print("Notes from the author:")
    print("All of the answer keys were automatically generated to reduce")
    print("input error from me. Still, I'm sure there are quite a few")
    print("errors hidden around here. You can always manually edit the")
    print("verb entries in the 'verbs.xlsx' file.")
    print("I used the masculine form for Indicative Passive, Perfect, and")
    print("Future Perfect. I don't think it effects memorization that much")
    print("(and I'm lazy)")
    print("but if you feel strongly that any gender should be available")
    print("contact me and I'll try to make it happen")
    print()
    print("+=============================================================+")
    input("Press 'Enter' to return to the main menu")

# this funtion adds verbs alphabetically to verb_list.xlsx
def set_verbs():
    from openpyxl import load_workbook
    FILE_NAME = 'verb_list.xlsx'
    wb = load_workbook(FILE_NAME, read_only = False)

    cls()

    end = 'y'
    while end != 'n':
        print("+=============================================================+")
        print('What conjugation is the verb?')
        print(wb.sheetnames)
        conj = input()

        ws = wb[conj]
        num_verbs = 0
        for x in ws.values:
            num_verbs = num_verbs + 1
        num_verbs = num_verbs - 1

        part = input("First principle part: ")
        part = part.lower()

        found = False
        i = 1;
        while (not found) and (i < num_verbs):
            i = i + 1
            if compare(part, ws.cell(row = i, column = 1).value) == 2:
                found = True
                print(part, "already existis in the verb list")
            elif compare(part, ws.cell(row = i, column = 1).value) == 1:
                found = True
                ws.insert_rows(i)
                ws.cell(row = i, column = 1, value = part)
                ws.cell(row = i, column = 2,
                        value = input("Second principle part: "))
                ws.cell(row = i, column = 3,
                        value = input("Third principle part: "))
                ws.cell(row = i, column = 4,
                        value = input("Fourth principle part: "))
                ws.cell(row = i, column = 5,
                        value = input("Translation: "))

        wb.save("verb_list.xlsx")
        end = input("Input another? y/n: ")

    wb.close()

# this funtion compares two verbs including macrons
# returns 0 if a > b
# returns 1 if a < b
# returns 2 if a = b
def compare(a, b):
    if a == b:
        return 2
    else :
        a = remove_macron(a)
        b = remove_macron(b)
        if a > b:
            return 0
        else:
            return 1

# this helper function returns a given string without macrons
def remove_macron(latin):
    non_latin = ""
    for i in range(len(latin)):
        if ord(latin[i]) == 257:
            non_latin = non_latin + 'a'
        elif ord(latin[i]) == 275:
            non_latin = non_latin + 'e'
        elif ord(latin[i]) == 299:
            non_latin = non_latin + 'i'
        elif ord(latin[i]) == 333:
            non_latin = non_latin + 'o'
        elif ord(latin[i]) == 363:
            non_latin = non_latin + 'u'
        else:
            non_latin = non_latin + latin[i]

    return non_latin

        
            

# this function copies all verbs from an excel file
# and generates the answer key in another
def copy_verbs():
    from openpyxl import load_workbook
    copy_from = 'verb_list.xlsx'
    copy_to = 'verbs.xlsx'
    wb_copy_from = load_workbook(copy_from, read_only = True)
    wb_copy_to = load_workbook(copy_to, read_only = False)

    cls()
    print("+=============================================================+")
    a = input("Enable verbose mode? y/n ")
    if a == "y":
        verbose = True
    else:
        verbose = False

    for ws_copy_from in wb_copy_from:
        conj = ws_copy_from.title

        ws_copy_to = wb_copy_to[conj]

        numRows = 0
        for x in ws_copy_from.values: # count the number of verbs to copy
            numRows = numRows + 1

        if verbose:
            print("copying ", numRows - 1, "in the ", conj, "conjugation")

        base = ["a", "a", "a", "a"]
        for j in range(numRows - 1): # iterate through every verb
            for x in range(4): # iterate through each principle part
                part = ws_copy_from.cell(row = j + 2, column = x + 1).value
                ws_copy_to.cell(row = j + 2, column = x + 1, value = part)

                # strip the principle parts of their first person endings
                if x == 0 or x == 2:
                    base[x] = part[:-1]
                elif x == 1:
                    base[x] = part[:-3]
                else:
                    base[x] = part[:-2]

            ws_copy_to.cell(row = j + 2, column = 5, value =
                            ws_copy_from.cell(row = j + 2, column = 5).value)

            end_file = open(conj + '.txt', "r", encoding='utf-8')
            pointer = int(5)
            while pointer < 113: # iterate through every verb form
                pointer = pointer + 1
                index = int((pointer - 6) / 6)
                if index >= 13 and index <= 15: #fourth part
                    root = base[3]

                elif (index >= 3 and index <=5) or index == 8 or index == 9: #third part
                    root = base[2]

                else: #first part
                    root = base[1]

                # add the ending to the principle part
                c = ws_copy_to.cell(row = j + 2, column = pointer,
                            value = root + end_file.readline().strip())

            end_file.close
        if verbose:
            print("\ncopy to ", conj, "conjugation complete\n\n")
        
    wb_copy_to.save(copy_to)
    print("Done")
    wb_copy_from.close()
    wb_copy_to.close()

# main
while True:
    cls()
    print("+=============================================================+")
    print("Welcome to simple latin verb practise!")
    print("Pick one of the options below to get started:")
    print("1) Start practising")
    print("2) How To guide")
    print("3) Add vocab")
    print("4) Generate new answer key")
    print("5) Exit")

    try:
        choice = int(input())
    except:
        choice = 6
        
    if choice == 1:
        verbs()
    elif choice == 2:
        how_to()
    elif choice == 3:
        set_verbs()
    elif choice == 4:
        copy_verbs()
    elif choice == 5:
        print("Goodbye!")
        break;
    else:
        print("Opps! wrong input")
